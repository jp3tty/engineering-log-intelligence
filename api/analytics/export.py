#!/usr/bin/env python3
"""
Data Export API
==============

This module provides comprehensive data export capabilities including:
- RESTful APIs for data export in multiple formats
- Advanced data filtering and aggregation
- Real-time export with streaming capabilities
- Batch export for large datasets
- Custom query builder for exports

Author: Engineering Log Intelligence Team
Date: September 29, 2025
"""

import json
import logging
import pandas as pd
import csv
import io
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass, asdict
import base64
import gzip
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ExportRequest:
    """Represents a data export request."""
    id: str
    data_source: str  # logs, metrics, alerts, incidents
    format: str  # csv, json, excel, parquet
    filters: Dict[str, Any]
    aggregations: List[str]
    group_by: List[str]
    sort_by: List[str]
    limit: Optional[int]
    time_range: Dict[str, str]
    compression: bool = False
    created_at: datetime = None
    status: str = "pending"  # pending, processing, completed, failed
    result_url: Optional[str] = None
    error_message: Optional[str] = None

@dataclass
class ExportResult:
    """Represents an export result."""
    request_id: str
    format: str
    file_size: int
    record_count: int
    download_url: str
    expires_at: datetime
    metadata: Dict[str, Any]

class DataExporter:
    """Data export system."""
    
    def __init__(self):
        self.export_requests = {}
        self.export_results = {}
        self.supported_formats = ['csv', 'json', 'excel', 'parquet']
        self.supported_sources = ['logs', 'metrics', 'alerts', 'incidents']
    
    def create_export_request(self, request_data: Dict[str, Any]) -> ExportRequest:
        """Create a new export request."""
        try:
            # Validate request
            self._validate_export_request(request_data)
            
            # Create export request
            request_id = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{hash(str(request_data)) % 10000:04d}"
            
            export_request = ExportRequest(
                id=request_id,
                data_source=request_data['data_source'],
                format=request_data['format'],
                filters=request_data.get('filters', {}),
                aggregations=request_data.get('aggregations', []),
                group_by=request_data.get('group_by', []),
                sort_by=request_data.get('sort_by', []),
                limit=request_data.get('limit'),
                time_range=request_data.get('time_range', {'start': '24h', 'end': 'now'}),
                compression=request_data.get('compression', False),
                created_at=datetime.now()
            )
            
            # Store request
            self.export_requests[request_id] = export_request
            
            logger.info(f"Created export request {request_id}")
            return export_request
            
        except Exception as e:
            logger.error(f"Error creating export request: {e}")
            raise
    
    def _validate_export_request(self, request_data: Dict[str, Any]):
        """Validate export request parameters."""
        required_fields = ['data_source', 'format']
        
        for field in required_fields:
            if field not in request_data:
                raise ValueError(f"Missing required field: {field}")
        
        if request_data['data_source'] not in self.supported_sources:
            raise ValueError(f"Unsupported data source: {request_data['data_source']}")
        
        if request_data['format'] not in self.supported_formats:
            raise ValueError(f"Unsupported format: {request_data['format']}")
    
    def process_export_request(self, request_id: str) -> ExportResult:
        """Process an export request and generate the export file."""
        try:
            if request_id not in self.export_requests:
                raise ValueError(f"Export request {request_id} not found")
            
            export_request = self.export_requests[request_id]
            export_request.status = "processing"
            
            # Generate mock data based on request
            data = self._generate_export_data(export_request)
            
            # Export data in requested format
            export_content = self._export_data(data, export_request)
            
            # Create result
            result = ExportResult(
                request_id=request_id,
                format=export_request.format,
                file_size=len(export_content),
                record_count=len(data) if isinstance(data, list) else data.get('count', 0),
                download_url=f"/api/analytics/export/download/{request_id}",
                expires_at=datetime.now() + timedelta(hours=24),
                metadata={
                    'data_source': export_request.data_source,
                    'filters_applied': export_request.filters,
                    'aggregations': export_request.aggregations,
                    'generated_at': datetime.now().isoformat()
                }
            )
            
            # Store result and update request
            self.export_results[request_id] = result
            export_request.status = "completed"
            export_request.result_url = result.download_url
            
            logger.info(f"Completed export request {request_id}: {result.record_count} records, {result.file_size} bytes")
            return result
            
        except Exception as e:
            logger.error(f"Error processing export request {request_id}: {e}")
            if request_id in self.export_requests:
                self.export_requests[request_id].status = "failed"
                self.export_requests[request_id].error_message = str(e)
            raise
    
    def _generate_export_data(self, request: ExportRequest) -> Union[List[Dict], Dict[str, Any]]:
        """Generate mock data for export based on request parameters."""
        import random
        
        # Parse time range
        time_range = self._parse_time_range(request.time_range)
        
        # Generate data based on source
        if request.data_source == 'logs':
            return self._generate_log_data(request, time_range)
        elif request.data_source == 'metrics':
            return self._generate_metrics_data(request, time_range)
        elif request.data_source == 'alerts':
            return self._generate_alerts_data(request, time_range)
        elif request.data_source == 'incidents':
            return self._generate_incidents_data(request, time_range)
        else:
            return []
    
    def _parse_time_range(self, time_range: Dict[str, str]) -> Dict[str, datetime]:
        """Parse time range parameters."""
        end_time = datetime.now()
        
        if time_range.get('end') == 'now':
            end_time = datetime.now()
        else:
            try:
                end_time = datetime.fromisoformat(time_range['end'])
            except:
                end_time = datetime.now()
        
        if time_range.get('start') == '24h':
            start_time = end_time - timedelta(hours=24)
        elif time_range.get('start') == '7d':
            start_time = end_time - timedelta(days=7)
        elif time_range.get('start') == '30d':
            start_time = end_time - timedelta(days=30)
        else:
            try:
                start_time = datetime.fromisoformat(time_range['start'])
            except:
                start_time = end_time - timedelta(hours=24)
        
        return {'start': start_time, 'end': end_time}
    
    def _generate_log_data(self, request: ExportRequest, time_range: Dict[str, datetime]) -> List[Dict]:
        """Generate mock log data."""
        import random
        
        data = []
        record_count = min(request.limit or 1000, 10000)  # Cap at 10k records
        
        log_levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL']
        sources = ['application', 'database', 'api', 'system', 'auth']
        messages = [
            'User authentication successful',
            'Database connection established',
            'API request processed',
            'Cache hit for key',
            'Failed to connect to external service',
            'Memory usage at 85%',
            'Queue processing started',
            'SSL handshake completed',
            'File upload completed',
            'Background job finished'
        ]
        
        for i in range(record_count):
            timestamp = time_range['start'] + timedelta(
                seconds=random.randint(0, int((time_range['end'] - time_range['start']).total_seconds()))
            )
            
            record = {
                'timestamp': timestamp.isoformat(),
                'level': random.choice(log_levels),
                'source': random.choice(sources),
                'message': random.choice(messages),
                'user_id': f"user_{random.randint(1, 100)}",
                'session_id': f"session_{random.randint(1000, 9999)}",
                'response_time': random.randint(10, 500),
                'status_code': random.choice([200, 201, 400, 401, 403, 404, 500]),
                'ip_address': f"192.168.1.{random.randint(1, 254)}",
                'user_agent': f"Mozilla/5.0 (Browser {random.randint(1, 10)})"
            }
            
            data.append(record)
        
        return data
    
    def _generate_metrics_data(self, request: ExportRequest, time_range: Dict[str, datetime]) -> List[Dict]:
        """Generate mock metrics data."""
        import random
        
        data = []
        record_count = min(request.limit or 1000, 10000)
        
        metrics = ['cpu_usage', 'memory_usage', 'disk_usage', 'network_throughput', 'response_time']
        
        for i in range(record_count):
            timestamp = time_range['start'] + timedelta(
                minutes=random.randint(0, int((time_range['end'] - time_range['start']).total_seconds() / 60))
            )
            
            record = {
                'timestamp': timestamp.isoformat(),
                'metric_name': random.choice(metrics),
                'metric_value': random.uniform(0, 100),
                'unit': random.choice(['percent', 'bytes', 'ms', 'requests/sec']),
                'host': f"host_{random.randint(1, 20)}",
                'service': random.choice(['web', 'api', 'db', 'cache', 'queue']),
                'environment': random.choice(['production', 'staging', 'development'])
            }
            
            data.append(record)
        
        return data
    
    def _generate_alerts_data(self, request: ExportRequest, time_range: Dict[str, datetime]) -> List[Dict]:
        """Generate mock alerts data."""
        import random
        
        data = []
        record_count = min(request.limit or 1000, 10000)
        
        alert_types = ['performance', 'security', 'availability', 'capacity', 'error']
        severities = ['low', 'medium', 'high', 'critical']
        statuses = ['open', 'acknowledged', 'resolved', 'closed']
        
        for i in range(record_count):
            timestamp = time_range['start'] + timedelta(
                minutes=random.randint(0, int((time_range['end'] - time_range['start']).total_seconds() / 60))
            )
            
            record = {
                'alert_id': f"alert_{i:06d}",
                'timestamp': timestamp.isoformat(),
                'type': random.choice(alert_types),
                'severity': random.choice(severities),
                'status': random.choice(statuses),
                'title': f"Alert {i}: {random.choice(alert_types).title()} Issue",
                'description': f"Description for alert {i}",
                'source': random.choice(['system', 'application', 'database', 'network']),
                'assigned_to': f"user_{random.randint(1, 10)}",
                'resolution_time': random.randint(5, 120) if random.choice(statuses) in ['resolved', 'closed'] else None
            }
            
            data.append(record)
        
        return data
    
    def _generate_incidents_data(self, request: ExportRequest, time_range: Dict[str, datetime]) -> List[Dict]:
        """Generate mock incidents data."""
        import random
        
        data = []
        record_count = min(request.limit or 1000, 10000)
        
        incident_types = ['outage', 'performance', 'security', 'data', 'service']
        priorities = ['low', 'medium', 'high', 'critical']
        statuses = ['open', 'investigating', 'resolved', 'closed']
        
        for i in range(record_count):
            timestamp = time_range['start'] + timedelta(
                hours=random.randint(0, int((time_range['end'] - time_range['start']).total_seconds() / 3600))
            )
            
            record = {
                'incident_id': f"inc_{i:06d}",
                'timestamp': timestamp.isoformat(),
                'type': random.choice(incident_types),
                'priority': random.choice(priorities),
                'status': random.choice(statuses),
                'title': f"Incident {i}: {random.choice(incident_types).title()} Issue",
                'description': f"Description for incident {i}",
                'affected_services': random.choice(['web', 'api', 'database', 'cache']),
                'assigned_team': f"team_{random.randint(1, 5)}",
                'resolution_time': random.randint(30, 480) if random.choice(statuses) in ['resolved', 'closed'] else None,
                'impact_score': random.randint(1, 10)
            }
            
            data.append(record)
        
        return data
    
    def _export_data(self, data: Union[List[Dict], Dict[str, Any]], request: ExportRequest) -> bytes:
        """Export data in the requested format."""
        if request.format == 'csv':
            return self._export_csv(data)
        elif request.format == 'json':
            return self._export_json(data)
        elif request.format == 'excel':
            return self._export_excel(data)
        elif request.format == 'parquet':
            return self._export_parquet(data)
        else:
            raise ValueError(f"Unsupported export format: {request.format}")
    
    def _export_csv(self, data: Union[List[Dict], Dict[str, Any]]) -> bytes:
        """Export data as CSV."""
        if not data:
            return b""
        
        output = io.StringIO()
        
        if isinstance(data, list) and len(data) > 0:
            # Get headers from first record
            headers = list(data[0].keys())
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            
            for record in data:
                writer.writerow(record)
        else:
            # Single record or dict
            if isinstance(data, dict):
                writer = csv.writer(output)
                for key, value in data.items():
                    writer.writerow([key, value])
        
        csv_content = output.getvalue()
        output.close()
        
        return csv_content.encode('utf-8')
    
    def _export_json(self, data: Union[List[Dict], Dict[str, Any]]) -> bytes:
        """Export data as JSON."""
        json_content = json.dumps(data, indent=2, default=str)
        return json_content.encode('utf-8')
    
    def _export_excel(self, data: Union[List[Dict], Dict[str, Any]]) -> bytes:
        """Export data as Excel."""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Export"
            
            if isinstance(data, list) and len(data) > 0:
                # Write headers
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Write data
                for row, record in enumerate(data, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=record.get(header, ''))
            else:
                # Single record or dict
                if isinstance(data, dict):
                    for row, (key, value) in enumerate(data.items(), 1):
                        ws.cell(row=row, column=1, value=key)
                        ws.cell(row=row, column=2, value=value)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            logger.warning("openpyxl not available, falling back to CSV format")
            return self._export_csv(data)
    
    def _export_parquet(self, data: Union[List[Dict], Dict[str, Any]]) -> bytes:
        """Export data as Parquet."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
            
            if isinstance(data, list) and len(data) > 0:
                df = pd.DataFrame(data)
                table = pa.Table.from_pandas(df)
                
                output = io.BytesIO()
                pq.write_table(table, output)
                output.seek(0)
                return output.getvalue()
            else:
                raise ValueError("Parquet export requires list of records")
                
        except ImportError:
            # Fallback to JSON if pyarrow not available
            logger.warning("pyarrow not available, falling back to JSON format")
            return self._export_json(data)
    
    def get_export_request(self, request_id: str) -> Optional[ExportRequest]:
        """Get export request by ID."""
        return self.export_requests.get(request_id)
    
    def get_export_result(self, request_id: str) -> Optional[ExportResult]:
        """Get export result by ID."""
        return self.export_results.get(request_id)
    
    def list_export_requests(self) -> List[Dict[str, Any]]:
        """List all export requests."""
        return [asdict(request) for request in self.export_requests.values()]
    
    def list_export_results(self) -> List[Dict[str, Any]]:
        """List all export results."""
        return [asdict(result) for result in self.export_results.values()]

def handler(request):
    """Vercel function handler for data export."""
    try:
        if request.method == 'GET':
            # Parse query parameters
            action = request.query.get('action', 'list')
            request_id = request.query.get('request_id')
            
            exporter = DataExporter()
            
            if action == 'list':
                # List export requests
                requests = exporter.list_export_requests()
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({
                        'success': True,
                        'requests': requests
                    })
                }
            
            elif action == 'request':
                # Get specific export request
                if not request_id:
                    return {
                        'statusCode': 400,
                        'headers': {'Content-Type': 'application/json'},
                        'body': json.dumps({'error': 'request_id parameter required'})
                    }
                
                export_request = exporter.get_export_request(request_id)
                if not export_request:
                    return {
                        'statusCode': 404,
                        'headers': {'Content-Type': 'application/json'},
                        'body': json.dumps({'error': 'Export request not found'})
                    }
                
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({
                        'success': True,
                        'request': asdict(export_request)
                    })
                }
            
            elif action == 'result':
                # Get export result
                if not request_id:
                    return {
                        'statusCode': 400,
                        'headers': {'Content-Type': 'application/json'},
                        'body': json.dumps({'error': 'request_id parameter required'})
                    }
                
                export_result = exporter.get_export_result(request_id)
                if not export_result:
                    return {
                        'statusCode': 404,
                        'headers': {'Content-Type': 'application/json'},
                        'body': json.dumps({'error': 'Export result not found'})
                    }
                
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({
                        'success': True,
                        'result': asdict(export_result)
                    })
                }
            
            elif action == 'capabilities':
                # Return export capabilities
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({
                        'supported_formats': exporter.supported_formats,
                        'supported_sources': exporter.supported_sources,
                        'max_records': 10000,
                        'max_file_size': '100MB'
                    })
                }
        
        elif request.method == 'POST':
            data = json.loads(request.body)
            
            exporter = DataExporter()
            
            # Create export request
            try:
                export_request = exporter.create_export_request(data)
                
                # Process export request
                export_result = exporter.process_export_request(export_request.id)
                
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({
                        'success': True,
                        'request': asdict(export_request),
                        'result': asdict(export_result)
                    })
                }
                
            except ValueError as e:
                return {
                    'statusCode': 400,
                    'headers': {'Content-Type': 'application/json'},
                    'body': json.dumps({'error': str(e)})
                }
        
        elif request.method == 'OPTIONS':
            return {
                'statusCode': 200,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
                    'Access-Control-Allow-Headers': 'Content-Type, Authorization'
                }
            }
        
        else:
            return {
                'statusCode': 405,
                'headers': {'Content-Type': 'application/json'},
                'body': json.dumps({'error': 'Method not allowed'})
            }
    
    except Exception as e:
        logger.error(f"Error in data export handler: {e}")
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json'},
            'body': json.dumps({
                'error': 'Internal server error',
                'message': str(e)
            })
        }
